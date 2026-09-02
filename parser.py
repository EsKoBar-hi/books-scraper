# -*- coding: utf-8 -*-
"""
Simple web scraper for the sandbox book catalog http://books.toscrape.com
Collects from the first page: book title, price, and stock status.
Saves the result to an Excel file (books.xlsx).
"""

import requests               # library for sending HTTP requests and downloading the page
from bs4 import BeautifulSoup # library for parsing HTML content
import pandas as pd           # library for working with tabular data and exporting to Excel


def get_books_data(url: str) -> list:
    """
    Downloads the HTML page at the given url, parses it,
    and returns a list of dictionaries with book data.
    """
    try:
        # send a GET request to the site; timeout=10 means wait max 10 seconds for a response
        response = requests.get(url, timeout=10)

        # force UTF-8 decoding so special characters (like £) don't get mangled
        response.encoding = "utf-8"

        # raise_for_status() throws an exception if the server returned an error code (404, 500, etc.)
        response.raise_for_status()

    except requests.exceptions.RequestException as error:
        # catches any network error: site not responding, no internet, timeout, etc.
        print(f"Error while requesting the site: {error}")
        return []  # return an empty list so the rest of the program doesn't crash

    # parse the downloaded HTML into a BeautifulSoup object
    # response.text is the page's text, "html.parser" is Python's built-in parser
    soup = BeautifulSoup(response.text, "html.parser")

    # every book on the page is wrapped in an <article class="product_pod"> tag
    # find_all() returns ALL matching tags as a list
    book_cards = soup.find_all("article", class_="product_pod")

    books_list = []  # will hold one dictionary per book

    # loop through every book card found on the page
    for card in book_cards:

        # the book title is stored in the "title" attribute of the <a> tag inside <h3>
        title = card.h3.a["title"]

        # price is inside <p class="price_color">; get_text(strip=True) removes extra whitespace
        price = card.find("p", class_="price_color").get_text(strip=True)

        # stock status is inside <p class="instock availability">
        availability = card.find("p", class_="instock availability").get_text(strip=True)

        # build a dictionary with this book's data
        book_info = {
            "title": title,           # book title
            "price": price,           # price (with currency symbol, as shown on the site)
            "in_stock": availability, # stock status ("In stock", etc.)
        }

        books_list.append(book_info)  # add this book to the overall list

    return books_list  # return the full list of collected books


def save_to_excel(data: list, filename: str = "books.xlsx") -> None:
    """
    Saves the collected data to an Excel file using pandas,
    then auto-adjusts column widths so text isn't cut off.
    """
    if not data:
        # nothing to save (e.g. because of a network error above)
        print("No data to save — file will not be created.")
        return

    # turn the list of dictionaries into a pandas table (DataFrame)
    df = pd.DataFrame(data)

    # index=False prevents pandas from adding an extra row-number column
    # engine="openpyxl" is required to write the .xlsx format
    df.to_excel(filename, index=False, engine="openpyxl")

    # --- auto-adjust column widths so the text is fully visible ---
    # openpyxl lets us reopen the saved file and tweak formatting
    from openpyxl import load_workbook

    workbook = load_workbook(filename)   # reopen the file we just saved
    worksheet = workbook.active          # get the active (first) sheet

    # loop through each column in the DataFrame to calculate a good width
    for column_index, column_name in enumerate(df.columns, start=1):
        # find the longest string in this column (including the header itself)
        max_length = max(
            df[column_name].astype(str).map(len).max(),  # longest value in the column
            len(str(column_name))                          # length of the header
        )
        # openpyxl uses column letters (A, B, C...) — get_column_letter converts the index
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(column_index)

        # set the width with a little extra padding for readability
        worksheet.column_dimensions[column_letter].width = max_length + 4

    workbook.save(filename)  # save the formatting changes back to the file

    print(f"Done! Data saved to: {filename}")


def main():
    """
    Main entry point of the script.
    """
    # URL of the first page of the book catalog
    target_url = "http://books.toscrape.com/index.html"

    print("Starting data collection...")

    # scrape the books from the target page
    books = get_books_data(target_url)

    print(f"Books collected: {len(books)}")

    # save the results to an Excel file
    save_to_excel(books, "books.xlsx")


# this ensures main() only runs when the script is executed directly,
# not when it's imported as a module in another script
if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Python web scraper for the sandbox book catalog http://books.toscrape.com
Collects title, price, and stock status from ALL pages of the catalog
(not just the first one) and saves the result to an Excel file.
"""

import requests               # library for sending HTTP requests and downloading pages
from bs4 import BeautifulSoup # library for parsing HTML content
import pandas as pd           # library for working with tabular data and exporting to Excel
import time                   # used to add a short delay between requests


def get_books_from_page(url: str) -> tuple:
    """
    Downloads a single catalog page and extracts book data from it.
    Returns a tuple: (list_of_books, url_of_next_page_or_None).
    """
    try:
        response = requests.get(url, timeout=10)
        response.encoding = "utf-8"  # force UTF-8 so special characters (like £) don't get mangled
        response.raise_for_status()

    except requests.exceptions.RequestException as error:
        print(f"Error while requesting {url}: {error}")
        return [], None  # on failure, return an empty page and stop pagination

    soup = BeautifulSoup(response.text, "html.parser")

    # every book on the page is wrapped in an <article class="product_pod"> tag
    book_cards = soup.find_all("article", class_="product_pod")

    books_on_this_page = []

    for card in book_cards:
        title = card.h3.a["title"]
        price = card.find("p", class_="price_color").get_text(strip=True)
        availability = card.find("p", class_="instock availability").get_text(strip=True)

        books_on_this_page.append({
            "title": title,
            "price": price,
            "in_stock": availability,
        })

    # find the "next" pagination link, if it exists
    next_button = soup.find("li", class_="next")

    if next_button:
        # the href is relative (e.g. "page-2.html"), so we build the full URL manually
        next_href = next_button.a["href"]
        # base_url is everything up to the last "/" in the current page's URL
        base_url = url.rsplit("/", 1)[0]
        next_page_url = f"{base_url}/{next_href}"
    else:
        next_page_url = None  # no more pages left

    return books_on_this_page, next_page_url


def get_all_books(start_url: str) -> list:
    """
    Loops through every page of the catalog (following the "next" links)
    and collects all books into a single list.
    """
    all_books = []
    current_url = start_url
    page_number = 1

    while current_url:
        print(f"Scraping page {page_number}...")

        books_on_page, next_url = get_books_from_page(current_url)
        all_books.extend(books_on_page)  # add this page's books to the overall list

        current_url = next_url
        page_number += 1

        # small delay between requests to avoid hammering the server
        time.sleep(0.5)

    return all_books


def save_to_excel(data: list, filename: str = "books.xlsx") -> None:
    """
    Saves the collected data to an Excel file using pandas,
    then auto-adjusts column widths so text isn't cut off.
    """
    if not data:
        print("No data to save — file will not be created.")
        return

    df = pd.DataFrame(data)
    df.to_excel(filename, index=False, engine="openpyxl")

    # auto-adjust column widths
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(filename)
    worksheet = workbook.active

    for column_index, column_name in enumerate(df.columns, start=1):
        max_length = max(
            df[column_name].astype(str).map(len).max(),
            len(str(column_name))
        )
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = max_length + 4

    workbook.save(filename)

    print(f"Done! Data saved to: {filename}")


def main():
    """
    Main entry point of the script.
    """
    start_url = "http://books.toscrape.com/index.html"

    print("Starting data collection from the entire catalog...")

    books = get_all_books(start_url)

    print(f"Total books collected: {len(books)}")

    save_to_excel(books, "books.xlsx")


if __name__ == "__main__":
    main()
